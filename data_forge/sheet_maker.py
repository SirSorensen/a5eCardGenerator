from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from data_forge.settings import *
from data_forge.data_structures.card import Card

class SheetMaker:
    def __init__(self, filepath : str = 'Cards.xlsm', read = False):
        self.wb = load_workbook(filepath, keep_vba=True, rich_text=True, data_only=read)
        self.ws = self.wb.active

    def add_card_sheet(self, sheet_name):
        self.ws = self.wb.create_sheet(sheet_name)
        self.wb.active = self.wb.index(self.ws)
        self.__save_card_titles()

    def add_row(self, row):
        self.ws.append(row)

    def save(self):
        self.wb.active = 0
        self.wb.save(r"Outputs/Sheets/Cards.xlsm")
    
    def __save_card_titles(self):
        fields = ["Title", "Name"]
        self.add_row(fields)
        self.__style_card_titles()

    
    def __style_card_titles(self):
        for row in self.ws.iter_rows(min_row=1, max_col=self.ws.max_column, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.style = "Accent1"

    def insert_card_properties(self, card : Card):
        field_values = [card.title, card.name]
        self.add_row(field_values)
    
    def read_input(self) -> list:
        result = []
        input_sheet = self.wb["Input"]
        for row in input_sheet.iter_rows(min_row=2, max_col=input_sheet.max_column, max_row=input_sheet.max_row):
            amount = row[0].value
            name = row[3].value

            if amount != None and name != None and amount > 0:
                result.append((row[0].value, row[3].value))
        
        return result


    def gen_name_list(self, amount):
        name_title = self.ws.title
        defined_name = DefinedName(name_title, attr_text=f"'{name_title}'!$A$2:$A${amount+1}")
        self.wb.defined_names[name_title] = defined_name   

